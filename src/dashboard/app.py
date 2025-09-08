"""Minimal Flask dashboard for uploading receipts and viewing costs.

Supports multi-file uploads and background OCR processing to handle
10–50 photos per session without blocking the request.
"""
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from werkzeug.utils import secure_filename

from pathlib import Path
import sys
import os
import json
from uuid import uuid4
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
from datetime import datetime
import io
import csv
from email.message import EmailMessage
import smtplib

# Optional Redis/RQ queue for OCR
try:
    from redis import Redis  # type: ignore
    from rq import Queue, Retry  # type: ignore
    from rq.job import Job  # type: ignore
    from rq.exceptions import NoSuchJobError  # type: ignore
except Exception:  # pragma: no cover
    Redis = None  # type: ignore
    Queue = None  # type: ignore
    Retry = None  # type: ignore
    Job = None  # type: ignore
    NoSuchJobError = Exception  # type: ignore

# Allow importing modules from the parent src directory
sys.path.append(str(Path(__file__).resolve().parents[1]))
from ocr.receipt_ocr import extract_receipt
try:
    # When running as a package (e.g., gunicorn)
    from .storage import (
        save_ticket,
        save_from_json_path,
        backfill_uploads,
        spend_by_month,
        spend_by_week,
        create_share,
        get_share,
        get_share_items,
        list_job_totals,
        spend_by_job,
        spend_by_category,
        items_for_month,
    )
    from .tasks import ocr_process
except Exception:
    # When running as a script: fallback to local import
    from storage import (
        save_ticket,
        save_from_json_path,
        backfill_uploads,
        spend_by_month,
        spend_by_week,
        create_share,
        get_share,
        get_share_items,
        list_job_totals,
        spend_by_job,
        spend_by_category,
        items_for_month,
    )
    from tasks import ocr_process

app = Flask(__name__)

# Store uploads within the dashboard package directory to avoid permission issues
UPLOAD_DIR = Path(__file__).resolve().parent / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Config
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "20"))
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp", ".gif"}

# Background OCR executor (Tesseract runs in a separate process per call,
# so threads are acceptable and simpler cross-platform)
OCR_ASYNC = os.getenv("OCR_ASYNC", "1") == "1"
OCR_WORKERS = max(1, int(os.getenv("OCR_WORKERS", "2")))
_EXECUTOR = ThreadPoolExecutor(max_workers=OCR_WORKERS) if OCR_ASYNC else None
_INFLIGHT: set[str] = set()
_INFLIGHT_LOCK = Lock()

# RQ setup when REDIS_URL provided
REDIS_URL = os.getenv("REDIS_URL")
_RQ: Queue | None = None
if REDIS_URL and Redis and Queue:  # type: ignore
    try:
        _RQ_CONN = Redis.from_url(REDIS_URL)  # type: ignore
        _RQ = Queue("ocr", connection=_RQ_CONN)  # type: ignore
    except Exception:
        _RQ = None


@app.after_request
def _security_headers(resp):
    """Add a few lightweight security headers suitable for a small app."""
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Referrer-Policy", "no-referrer")
    # Allow Bootstrap + Chart.js CDN and inline scripts used by templates
    csp = (
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "script-src 'self' https://cdn.jsdelivr.net 'unsafe-inline'; "
        "connect-src 'self'"
    )
    resp.headers.setdefault("Content-Security-Policy", csp)
    return resp


@app.route("/")
def index():
    # Load all JSONs, then collapse duplicates by original name
    candidates = []
    for f in UPLOAD_DIR.glob("*.json"):
        try:
            data = json.loads(f.read_text())
        except json.JSONDecodeError as exc:
            data = {"raw_text": f"JSON error: {exc}"}
        display_name = data.get("original_name", data.get("file", f.stem))
        mtime = 0.0
        try:
            mtime = f.stat().st_mtime
        except Exception:
            pass
        candidates.append({
            "id": f.stem,
            "name": display_name,
            "data": data,
            "mtime": mtime,
        })

    # Prefer completed+filled entries over processing/empty for the same original file
    grouped: dict[str, dict] = {}
    for c in candidates:
        data = c["data"]
        processing = data.get("status") == "processing"
        # Job name is not required for fuel receipts
        job_ok = bool(data.get("job_name")) or (data.get("category") == "fuel")
        total_ok = data.get("total") is not None
        score = (
            0 if processing else 1,                 # prefer not processing
            (1 if job_ok else 0) + (1 if total_ok else 0),  # prefer both fields present
            c["mtime"],                              # prefer newer
        )
        key = c["name"]
        prev = grouped.get(key)
        if not prev or score > prev["_score"]:
            grouped[key] = {**c, "_score": score}

    entries = [{k: v for k, v in item.items() if k != "_score"} for item in grouped.values()]

    # Build batch groups: latest batches first, entries ordered by batch_seq
    def _parse_batch_ts(bid: str | None) -> float:
        if not bid:
            return -1.0
        try:
            ts = bid.split("_", 1)[0]
            dt = datetime.strptime(ts, "%Y%m%d%H%M%S")
            return dt.timestamp()
        except Exception:
            return 0.0

    groups_map: dict[str | None, dict] = {}
    for e in entries:
        data = e.get("data", {})
        bid = data.get("batch_id")
        g = groups_map.setdefault(bid, {"id": bid, "items": [], "batch_total": data.get("batch_total")})
        # Keep the max batch_total if present
        if data.get("batch_total") and (not g.get("batch_total") or data.get("batch_total") > g.get("batch_total")):
            g["batch_total"] = data.get("batch_total")
        g["items"].append(e)

    # Order entries within group
    for g in groups_map.values():
        g["items"].sort(key=lambda x: (x.get("data", {}).get("batch_seq") or 0, x["name"]))
        # Precompute title
        if g["id"]:
            ts = _parse_batch_ts(g["id"]) or 0
            try:
                title = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                title = g["id"]
            g["title"] = title
        else:
            g["title"] = "Other"

    # Sort groups: latest timestamp first; None (Other) last
    groups = sorted(groups_map.values(), key=lambda g: (_parse_batch_ts(g["id"]) if g["id"] else -1.0), reverse=True)

    # Auto-scan for orphan images without JSON and schedule processing
    if OCR_ASYNC:
        _bootstrap_pending_jobs()

    # Backfill DB so saved tickets are queryable in the future
    try:
        backfill_uploads(UPLOAD_DIR)
    except Exception:
        pass

    any_processing = any(e["data"].get("status") == "processing" for e in entries)
    return render_template("index.html", groups=groups, any_processing=any_processing)


@app.route("/upload", methods=["POST"])
def upload():
    files = request.files.getlist("receipt") or []
    # Collect only valid image files
    valid: list[tuple] = []
    for file in files:
        if not file or file.filename == "":
            continue
        filename = secure_filename(file.filename)
        if Path(filename).suffix.lower() not in ALLOWED_EXT:
            continue
        valid.append((file, filename))
    if not valid:
        return redirect("/")

    # Tag this batch and assign sequence numbers
    batch_id = datetime.utcnow().strftime("%Y%m%d%H%M%S") + f"_{uuid4().hex[:6]}"
    batch_total = len(valid)

    for idx, (file, filename) in enumerate(valid, start=1):
        dest = UPLOAD_DIR / f"{uuid4().hex}_{filename}"
        file.save(dest)
        meta = {"batch_id": batch_id, "batch_seq": idx, "batch_total": batch_total, "original_name": filename}
        if OCR_ASYNC:
            _schedule_ocr(dest, meta)
        else:
            data = extract_receipt(dest)
            out = {**meta, **data}
            json_path = dest.with_suffix(".json")
            json_path.write_text(json.dumps(out, indent=2))
            # Persist to DB if any field present
            try:
                save_from_json_path(json_path)
            except Exception:
                pass
            if not out.get("job_name") or out.get("total") is None:
                return redirect(url_for("classify", name=json_path.stem))
    return redirect("/")


@app.route("/classify/<name>", methods=["GET", "POST"])
def classify(name: str):
    json_path = UPLOAD_DIR / f"{name}.json"
    if not json_path.exists():
        return redirect("/")
    try:
        data = json.loads(json_path.read_text())
    except json.JSONDecodeError:
        data = {}

    # Do not allow edits while processing; show form after completion
    if data.get("status") == "processing" and request.method == "GET":
        return redirect("/")

    if request.method == "POST":
        job_name = (request.form.get("job_name") or "").strip() or None
        total_str = (request.form.get("total") or "").strip()
        category_raw = (request.form.get("category") or "").strip().lower()
        category = "fuel" if category_raw in {"fuel", "gas", "gasoline", "petrol"} else None
        try:
            total_val = float(total_str.replace(",", "")) if total_str else None
        except ValueError:
            total_val = None

        # If marked as fuel, ignore any provided job name
        if category == "fuel":
            job_name = None

        data["job_name"] = job_name
        data["total"] = total_val
        if category:
            data["category"] = category
        else:
            data.pop("category", None)
        data.pop("status", None)
        json_path.write_text(json.dumps(data, indent=2))
        # Persist classification edits to DB
        try:
            save_ticket(
                id=name,
                file=data.get("file"),
                original_name=data.get("original_name"),
                job_name=job_name,
                total=total_val,
                batch_id=data.get("batch_id"),
                batch_seq=data.get("batch_seq"),
                category=category,
            )
        except Exception:
            pass
        return redirect("/")

    return render_template("classify.html", file=name, data=data)


@app.route("/reports")
def reports():
    # Ensure DB has latest entries
    try:
        backfill_uploads(UPLOAD_DIR)
    except Exception:
        pass
    monthly = []
    weekly = []
    try:
        monthly = spend_by_month()
        weekly = spend_by_week()
    except Exception:
        pass
    return render_template("reports.html", monthly=monthly, weekly=weekly)


@app.route("/share/new")
def share_new():
    """Create a temporary share link to view job/total data.

    Optional query param `ttl` specifies minutes until expiration (default 60).
    """
    try:
        ttl = int(request.args.get("ttl", "60"))
        ttl = max(1, min(ttl, 7 * 24 * 60))  # clamp to [1 minute, 7 days]
    except Exception:
        ttl = 60
    token = create_share(ttl_minutes=ttl)
    return redirect(url_for("share_view", token=token))


@app.route("/share/<token>")
def share_view(token: str):
    meta = get_share(token)
    if not meta:
        return render_template("share.html", token=token, expired=True, items=[], meta=None)
    # Ensure DB has latest entries from uploads
    try:
        backfill_uploads(UPLOAD_DIR)
    except Exception:
        pass
    items = get_share_items(token) or []
    return render_template("share.html", token=token, expired=False, items=items, meta=meta)


# JSON APIs for charts
@app.route("/api/spend/by_job")
def api_spend_by_job():
    month = request.args.get("month") or None
    try:
        data = spend_by_job(month=month)
    except Exception:
        data = []
    return jsonify(data)


@app.route("/api/spend/by_category")
def api_spend_by_category():
    month = request.args.get("month") or None
    try:
        data = spend_by_category(month=month)
    except Exception:
        data = []
    return jsonify(data)


def _dataset_for_type(dtype: str):
    dtype = (dtype or "monthly").lower()
    if dtype == "monthly":
        return ("monthly", ["month", "total", "fuel_total", "other_total"], spend_by_month())
    if dtype == "weekly":
        return ("weekly", ["year_week", "total", "fuel_total", "other_total"], spend_by_week())
    if dtype == "by_job":
        return ("by_job", ["job_name", "total", "count"], spend_by_job())
    if dtype == "by_category":
        return ("by_category", ["category", "total", "count"], spend_by_category())
    # default
    return ("monthly", ["month", "total", "fuel_total", "other_total"], spend_by_month())


@app.route("/reports/export.csv")
def reports_export_csv():
    dtype = request.args.get("type", "monthly")
    name, headers, rows = _dataset_for_type(dtype)
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(headers)
    for row in rows:
        w.writerow([row.get(h) for h in headers])
    output.seek(0)
    filename = f"{name}.csv"
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/reports/export.xlsx")
def reports_export_xlsx():
    dtype = request.args.get("type", "monthly")
    name, headers, rows = _dataset_for_type(dtype)
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception:
        return "openpyxl not installed", 500
    wb = Workbook()
    ws = wb.active
    ws.title = name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])

    # Improve sheet: freeze headers
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass

    # Add basic charts on both the data sheet and a dedicated chart sheet
    try:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference  # type: ignore
        from openpyxl.styles import Font  # type: ignore
        # Bold header row
        try:
            for c in range(1, len(headers) + 1):
                ws.cell(row=1, column=c).font = Font(bold=True)
        except Exception:
            pass

        chart_ws = wb.create_sheet(title=f"{name}_chart")
        n = len(rows)
        max_row = max(2, n + 1)  # ensure at least header+one row for chart frame

        if dtype == "monthly":
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=min(4, len(headers)), max_row=max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart = BarChart(); chart.title = "Monthly Totals"; chart.y_axis.title = "Amount"; chart.x_axis.title = "Month"
            chart.add_data(data_ref, titles_from_header=True); chart.set_categories(cats_ref)
            chart1 = BarChart(); chart1.title = chart.title; chart1.y_axis.title = chart.y_axis.title; chart1.x_axis.title = chart.x_axis.title
            chart1.add_data(data_ref, titles_from_header=True); chart1.set_categories(cats_ref)
            ws.add_chart(chart1, "F2"); chart_ws.add_chart(chart, "A1")
        elif dtype == "weekly":
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=min(4, len(headers)), max_row=max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart = LineChart(); chart.title = "Weekly Totals"; chart.y_axis.title = "Amount"; chart.x_axis.title = "Week"
            chart.add_data(data_ref, titles_from_header=True); chart.set_categories(cats_ref)
            chart1 = LineChart(); chart1.title = chart.title; chart1.y_axis.title = chart.y_axis.title; chart1.x_axis.title = chart.x_axis.title
            chart1.add_data(data_ref, titles_from_header=True); chart1.set_categories(cats_ref)
            ws.add_chart(chart1, "F2"); chart_ws.add_chart(chart, "A1")
        elif dtype == "by_job":
            max_col = min(3, len(headers))
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart = BarChart(); chart.title = "Spend by Job"; chart.y_axis.title = "Amount / Count"; chart.x_axis.title = "Job"
            chart.add_data(data_ref, titles_from_header=True); chart.set_categories(cats_ref)
            chart1 = BarChart(); chart1.title = chart.title; chart1.y_axis.title = chart.y_axis.title; chart1.x_axis.title = chart.x_axis.title
            chart1.add_data(data_ref, titles_from_header=True); chart1.set_categories(cats_ref)
            ws.add_chart(chart1, "F2"); chart_ws.add_chart(chart, "A1")
        elif dtype == "by_category":
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            pie = PieChart(); pie.title = "Spend by Category"; pie.add_data(data_ref, titles_from_header=True); pie.set_categories(cats_ref)
            pie1 = PieChart(); pie1.title = pie.title; pie1.add_data(data_ref, titles_from_header=True); pie1.set_categories(cats_ref)
            ws.add_chart(pie1, "F2"); chart_ws.add_chart(pie, "A1")
    except Exception:
        pass
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{name}.xlsx",
    )


@app.route("/reports/export_full.xlsx")
def reports_export_full_xlsx():
    """Export a single workbook containing all datasets and a Dashboard sheet with charts."""
    # Gather data
    monthly_rows = spend_by_month()
    weekly_rows = spend_by_week()
    by_job_rows = spend_by_job()
    by_cat_rows = spend_by_category()

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except Exception:
        return "openpyxl not installed", 500

    wb = Workbook()

    def add_sheet(name: str, headers: list[str], rows: list[dict]):
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
        try:
            ws.freeze_panes = "A2"
            for c in range(1, len(headers) + 1):
                ws.cell(row=1, column=c).font = Font(bold=True)
        except Exception:
            pass
        return ws

    # Remove default first sheet and rebuild in clear order
    try:
        del wb[wb.active.title]
    except Exception:
        pass

    ws_monthly = add_sheet("monthly", ["month", "total", "fuel_total", "other_total"], monthly_rows)
    ws_weekly = add_sheet("weekly", ["year_week", "total", "fuel_total", "other_total"], weekly_rows)
    ws_job = add_sheet("by_job", ["job_name", "total", "count"], by_job_rows)
    ws_cat = add_sheet("by_category", ["category", "total", "count"], by_cat_rows)

    dash = wb.create_sheet(title="Dashboard")
    try:
        dash["A1"] = "Dashboard"
        dash["A1"].font = Font(bold=True, size=14)
    except Exception:
        pass

    # Charts: always add chart frames; references include at least header+one row
    try:
        # Monthly chart (bar) at A3
        n = len(monthly_rows); max_row = max(2, n + 1)
        data_ref = Reference(ws_monthly, min_col=2, min_row=1, max_col=4, max_row=max_row)
        cats_ref = Reference(ws_monthly, min_col=1, min_row=2, max_row=max_row)
        c = BarChart(); c.title = "Monthly Totals"; c.y_axis.title = "Amount"; c.x_axis.title = "Month"; c.width = 22; c.height = 13
        c.add_data(data_ref, titles_from_header=True); c.set_categories(cats_ref)
        dash.add_chart(c, "A3")

        # Weekly chart (line) at A22
        n = len(weekly_rows); max_row = max(2, n + 1)
        data_ref = Reference(ws_weekly, min_col=2, min_row=1, max_col=4, max_row=max_row)
        cats_ref = Reference(ws_weekly, min_col=1, min_row=2, max_row=max_row)
        c = LineChart(); c.title = "Weekly Totals"; c.y_axis.title = "Amount"; c.x_axis.title = "Week"; c.width = 22; c.height = 13
        c.add_data(data_ref, titles_from_header=True); c.set_categories(cats_ref)
        dash.add_chart(c, "A22")

        # Category pie at H3
        n = len(by_cat_rows); max_row = max(2, n + 1)
        data_ref = Reference(ws_cat, min_col=2, min_row=1, max_col=2, max_row=max_row)
        cats_ref = Reference(ws_cat, min_col=1, min_row=2, max_row=max_row)
        p = PieChart(); p.title = "Spend by Category"; p.add_data(data_ref, titles_from_header=True); p.set_categories(cats_ref); p.width = 18; p.height = 12
        dash.add_chart(p, "H3")

        # Top jobs bar at H22 (limit to top 10 rows)
        n = len(by_job_rows); n10 = min(n, 10); max_rows = max(2, n10 + 1)
        data_ref = Reference(ws_job, min_col=2, min_row=1, max_col=3, max_row=max_rows)
        cats_ref = Reference(ws_job, min_col=1, min_row=2, max_row=max_rows)
        c = BarChart(); c.type = "bar"; c.title = "Top Jobs (by Total)"; c.y_axis.title = "Job"; c.x_axis.title = "Amount / Count"; c.width = 22; c.height = 13
        c.add_data(data_ref, titles_from_header=True); c.set_categories(cats_ref)
        dash.add_chart(c, "H22")
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="full_report.xlsx",
    )


def _current_month_str() -> str:
    return datetime.utcnow().strftime("%Y-%m")


def _render_monthly_report_html(month: str) -> str:
    # Prepare data
    by_month = spend_by_month()
    by_job = spend_by_job(month=month)
    by_cat = spend_by_category(month=month)
    items = items_for_month(month)
    total = sum((r.get("total") or 0) for r in by_job)
    return render_template(
        "monthly_report.html",
        month=month,
        total=total,
        monthly=by_month,
        by_job=by_job,
        by_category=by_cat,
        items=items,
    )


@app.route("/reports/monthly.pdf")
def reports_monthly_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        return "WeasyPrint not installed", 500
    month = request.args.get("month") or _current_month_str()
    html = _render_monthly_report_html(month)
    pdf = HTML(string=html).write_pdf()
    return send_file(
        io.BytesIO(pdf),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"monthly_report_{month}.pdf",
    )


def _send_monthly_report_email(month: str) -> tuple[bool, str]:
    host = os.getenv("SMTP_HOST")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASS")
    sender = os.getenv("EMAIL_FROM") or (user or "")
    recipients = [e.strip() for e in (os.getenv("EMAIL_TO") or "").split(",") if e.strip()]
    starttls = os.getenv("SMTP_STARTTLS", "1") == "1"
    if not (host and sender and recipients):
        return False, "Missing SMTP_HOST/EMAIL_FROM/EMAIL_TO"
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        return False, "WeasyPrint not installed"
    html = _render_monthly_report_html(month)
    pdf_bytes = HTML(string=html).write_pdf()
    msg = EmailMessage()
    msg["Subject"] = f"Monthly Report {month}"
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.set_content(f"Attached: Monthly Report {month}")
    msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=f"monthly_report_{month}.pdf")
    try:
        with smtplib.SMTP(host, port) as smtp:
            smtp.ehlo()
            if starttls:
                smtp.starttls()
            if user and password:
                smtp.login(user, password)
            smtp.send_message(msg)
        return True, "sent"
    except Exception as e:
        return False, str(e)


@app.route("/reports/monthly/email")
def reports_monthly_email():
    month = request.args.get("month") or _current_month_str()
    ok, msg = _send_monthly_report_email(month)
    code = 200 if ok else 500
    return {"ok": ok, "message": msg, "month": month}, code


def _maybe_schedule_email():
    if os.getenv("SCHEDULE_EMAIL", "0") != "1":
        return
    try:
        from apscheduler.schedulers.background import BackgroundScheduler  # type: ignore
        from apscheduler.triggers.cron import CronTrigger  # type: ignore
    except Exception:
        return
    # Only start one scheduler per process; users should run one worker to avoid duplicates
    sched = BackgroundScheduler(timezone="UTC")
    # Daily at 08:00 UTC by default; override with EMAIL_CRON (e.g., "0 8 * * *")
    cron = os.getenv("EMAIL_CRON")
    if cron:
        try:
            trigger = CronTrigger.from_crontab(cron)
        except Exception:
            trigger = CronTrigger(hour=8, minute=0)
    else:
        trigger = CronTrigger(hour=8, minute=0)

    def _job():
        try:
            _send_monthly_report_email(_current_month_str())
        except Exception:
            pass

    sched.add_job(_job, trigger, id="monthly_email", replace_existing=True)
    sched.start()


# Initialize scheduler if configured
_maybe_schedule_email()


def _schedule_ocr(image_path: Path, meta: dict | None = None) -> None:
    """Create a placeholder JSON and schedule background OCR."""
    json_path = image_path.with_suffix(".json")
    # Derive original filename (after UUID_ prefix)
    n = image_path.name
    original_name = n.split("_", 1)[1] if "_" in n else n
    placeholder = {
        "status": "processing",
        "file": image_path.name,
        "original_name": original_name,
        "started_at": datetime.utcnow().isoformat() + "Z",
    }
    if meta:
        placeholder.update({k: v for k, v in meta.items() if k in {"batch_id", "batch_seq", "batch_total", "original_name"}})
    json_path.write_text(json.dumps(placeholder, indent=2))

    # Prefer Redis/RQ if configured
    if _RQ is not None:
        job_id = f"ocr:{image_path.name}"
        try:
            j = Job.fetch(job_id, connection=_RQ.connection)  # type: ignore
            status = j.get_status(refresh=True)
            if status in ("queued", "started", "deferred"):
                return
        except NoSuchJobError:
            pass
        except Exception:
            pass
        # Enqueue with retries and timeout
        try:
            _RQ.enqueue(
                ocr_process,
                str(image_path),
                job_id=job_id,
                retry=Retry(max=3, interval=[10, 60, 180]),  # type: ignore
                job_timeout=int(os.getenv("OCR_JOB_TIMEOUT", "600")),
                result_ttl=int(os.getenv("OCR_RESULT_TTL", "86400")),
                failure_ttl=int(os.getenv("OCR_FAILURE_TTL", "604800")),
            )
            return
        except Exception:
            # Fall back to local thread if enqueue fails
            pass

    # Fallback: local thread pool
    assert _EXECUTOR is not None
    key = str(image_path)
    with _INFLIGHT_LOCK:
        if key in _INFLIGHT:
            return
        _INFLIGHT.add(key)
    _EXECUTOR.submit(_process_and_write, image_path)


def _process_and_write(image_path: Path) -> None:
    key = str(image_path)
    try:
        data = extract_receipt(image_path)
        # Clear processing status on success
        data.pop("status", None)
        n = image_path.name
        original_name = n.split("_", 1)[1] if "_" in n else n
        # Carry over batch metadata from placeholder if present
        meta: dict = {}
        try:
            prev = json.loads(image_path.with_suffix(".json").read_text())
            for k in ("batch_id", "batch_seq", "batch_total"):
                if k in prev:
                    meta[k] = prev[k]
        except Exception:
            pass
        out = {"file": image_path.name, "original_name": original_name, **meta, **data}
        json_p = image_path.with_suffix(".json")
        json_p.write_text(json.dumps(out, indent=2))
        # Persist to DB
        try:
            save_from_json_path(json_p)
        except Exception:
            pass
    except Exception as exc:
        error = {"status": "error", "error": str(exc), "file": image_path.name}
        image_path.with_suffix(".json").write_text(json.dumps(error, indent=2))
    finally:
        with _INFLIGHT_LOCK:
            _INFLIGHT.discard(key)


def _bootstrap_pending_jobs() -> None:
    """On each index view, (re)schedule any images missing results."""
    # RQ mode: ensure missing/abandoned jobs are enqueued
    if _RQ is not None:
        for img in UPLOAD_DIR.iterdir():
            if img.is_file() and img.suffix.lower() in ALLOWED_EXT:
                jpath = img.with_suffix(".json")
                if not jpath.exists():
                    _schedule_ocr(img)
                    continue
                try:
                    meta = json.loads(jpath.read_text())
                except Exception:
                    continue
                if meta.get("status") == "processing":
                    job_id = f"ocr:{img.name}"
                    try:
                        j = Job.fetch(job_id, connection=_RQ.connection)  # type: ignore
                        status = j.get_status(refresh=True)
                        if status in ("queued", "started", "deferred"):
                            continue
                    except NoSuchJobError:
                        pass
                    except Exception:
                        pass
                    # (Re)enqueue
                    try:
                        _RQ.enqueue(
                            ocr_process,
                            str(img),
                            job_id=job_id,
                            retry=Retry(max=3, interval=[10, 60, 180]),  # type: ignore
                            job_timeout=int(os.getenv("OCR_JOB_TIMEOUT", "600")),
                        )
                    except Exception:
                        pass
        return

    # Thread fallback
    assert _EXECUTOR is not None
    for img in UPLOAD_DIR.iterdir():
        if img.is_file() and img.suffix.lower() in ALLOWED_EXT:
            j = img.with_suffix(".json")
            if not j.exists():
                _schedule_ocr(img)
                continue
            try:
                meta = json.loads(j.read_text())
            except Exception:
                continue
            if meta.get("status") == "processing":
                # Re-enqueue if previous run crashed, but avoid duplicates in this process
                with _INFLIGHT_LOCK:
                    if str(img) in _INFLIGHT:
                        continue
                    _INFLIGHT.add(str(img))
                _EXECUTOR.submit(_process_and_write, img)


if __name__ == "__main__":
    # Disable debug mode and reloader for lower resource use on small devices
    app.run(host="0.0.0.0", use_reloader=False)
